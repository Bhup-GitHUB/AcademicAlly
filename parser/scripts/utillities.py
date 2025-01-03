from openpyxl.utils import get_column_letter

def get_merged_cells(sheet):
    merged_cells_map = {}
    min_rows_set = set()
    merged_cells = sheet.merged_cells.ranges
    for merged_range in merged_cells:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if min_col <= 4 <= max_col:
            top_left_cell = sheet.cell(row=min_col, column=min_row)
            cell_value = top_left_cell.value
            merged_cells_map[cell_value] = (min_row, max_row)
            min_rows_set.add(min_row)
    return merged_cells_map, min_rows_set


def remove_whitespace(s):
    if s is None:
        return None
    return ''.join(s.split())


def return_day(row_val):
    if 8 <= row_val <= 35:
        return "Monday"
    elif 36 <= row_val <= 63:
        return "Tuesday"
    elif 64 <= row_val <= 91:
        return "Wednesday"
    elif 92 <= row_val <= 119:
        return "Thursday"
    elif 120 <= row_val <= 147:
        return "Friday"
    elif 148 <= row_val <= 175:
        return "Saturday"
    else:
        return None


def getTypeOfClass(s):
    for char in reversed(s):
        if char != ' ':
            return char
    return None


def get_formatted_time(sheet, row, column):
    cell_value = sheet.cell(row=row, column=column).value

    if not isinstance(cell_value, str):
        try:
            cell_value = cell_value.strftime("%I:%M %p")
        except AttributeError:
            cell_value = str(cell_value)

    return cell_value.strip()



def HandlePractical(sheet, x, y, subgroup, day, time, cell, result,room):
    # room = sheet.cell(row=x + 1, column=y).value + ' ' + sheet.cell(row=x + 2, column=y).value
    result[subgroup][day][time] = [cell, room]
    print(f"result[{subgroup}][{day}][{time}] = {cell}")

    # time2 = sheet.cell(row=x + 2, column=4).value.strftime("%I:%M %p")
    time2 = get_formatted_time(sheet, x + 2, 4)

    result[subgroup][day][time2] = [cell, room]
    print(f"result[{subgroup}][{day}][{time2}] = {cell}")
    skip_next = True


def HandleTutorial(sheet, x, y, subgroup, day, time, cell, result,room):
    tVal = None
    checker = sheet.cell(row=x + 2, column=y).value
    if (checker is not None and len(checker) <= 5):
        # room = sheet.cell(row=x + 1, column=y).value + ' ' + sheet.cell(row=x + 2, column=y).value
        time2 = get_formatted_time(sheet, x + 2, 4)
        result[subgroup][day][time2] = [cell, room]
        print(f"result[{subgroup}][{day}][{time2}] = {cell}")
        tVal = 1

    # room = sheet.cell(row=x + 1, column=y).value
    result[subgroup][day][time] = [cell, room]
    print(f"result[{subgroup}][{day}][{time}] = {cell}")
    return tVal

def count_merged_cells_along_row(ws, row, column):
    cell_address = f"{get_column_letter(column)}{row}"
    cell = ws[cell_address]
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:

            min_col, min_row, max_col, max_row = merged_range.bounds
            # if min_row == max_row:  # Ensure the merge is along a single row
            number_of_merged_cells_along_row = max_col - min_col + 1
            return number_of_merged_cells_along_row
    return 0
    #         else:
    #             return f'The cell at row {row}, column {column} is merged, but not along a single row.'
    # return f'The cell at row {row}, column {column} is not part of any merged range.'


def HandleLecture(sheet, x, y, subgroup, day, time, cell, result, merged_cells_map,room):
    # room = sheet.cell(row=x + 1, column=y).value
    lectureGroup = sheet.cell(row=4, column=y).value
    print(f"lectureGroup: {lectureGroup}")
    result[subgroup][day][time] = [cell, room]
    print(f"result[{subgroup}][{day}][{time}] = {cell}")

    # try:
    #     rangeVal = merged_cells_map[lectureGroup]
    # except KeyError:
    #     result[subgroup][day][time] = [cell, room]
    #     print(f"result[{subgroup}][{day}][{time}] = {cell}")
    #     return
    #
    # print(f"rangeVal: {rangeVal}")
    #
    # for m in range(y, rangeVal[1] + 1, 2):
    #     subgroup = remove_whitespace(sheet.cell(row=7, column=m).value)
    #     print(f"Processing subgroup at column={m}: {subgroup}")
    #     result[subgroup][day][time] = [cell, room]
    #     print(f"result[{subgroup}][{day}][{time}] = {cell}")


def parser(sheet, result):
    merged_cells_map = get_merged_cells(sheet)[0]

    row_size = sheet.max_row
    col_size = sheet.max_column

    skip_next = False
    for y in range(5, col_size, 2):
        for x in range(8, row_size, 2):

            if skip_next:
                skip_next = False
                continue

            subgroup = remove_whitespace(sheet.cell(row=5, column=y).value)

            day = return_day(x)
            if (day is None):
                break

            time = get_formatted_time(sheet, x, 4)

            cell = remove_whitespace(sheet.cell(row=x, column=y).value)

            # Debug prints
            print(
                f"Processing cell at row={x}, column={y} | subgroup: {subgroup} | time: {time} | day: {day} |  cell value: {cell}")


            if cell is not None:
                type = getTypeOfClass(cell)
                print(f"type of class: {type}")
                if type == 'L' or type =='T':
                    room = room = sheet.cell(row=x + 1, column=y).value
                elif type == 'P':
                    room = sheet.cell(row=x + 1, column=y).value + ' ' + sheet.cell(row=x + 2, column=y).value
                    print(room)


                merged_cells_count = count_merged_cells_along_row(sheet, x, y)
                for m in range(y, y + merged_cells_count, 2):
                    subgroup = remove_whitespace(sheet.cell(row=5, column=m).value)

                    if type == 'L':
                        HandleLecture(sheet, x, y, subgroup, day, time, cell, result, merged_cells_map, room)

                    elif type == 'P':
                        HandlePractical(sheet, x, y, subgroup, day, time, cell, result, room)
                        skip_next = True

                    elif type == 'T':
                        tVal = HandleTutorial(sheet, x, y, subgroup, day, time, cell, result, room)
                        if tVal is not None:
                            skip_next = True

